#from isort import file
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.chart.layout import Layout, ManualLayout
from pythonLibs.fileHandler.FileHandlerAscii import *
import re

EXCEL_COLOR_LIST = ['ltSteelBlue', 'dkGoldenrod', 'firebrick', 'lightSlateGray', 'ltSlateGray', 'steelBlue', 'bisque', 'crimson', 'darkViolet', 'lime', 'darkOrange', 'floralWhite', 'lightCoral', 'olive', 'dkViolet', 'tomato', 'hotPink', 'darkGrey', 'slateBlue', 'mediumTurquoise', 'paleGreen', 'dkBlue', 'black', 'lawnGreen', 'limeGreen', 'paleGoldenrod', 'darkSalmon', 'medSlateBlue', 'lightGray', 'dkSlateGrey', 'dodgerBlue', 'peachPuff', 'darkOliveGreen', 'dkSeaGreen', 'lightGrey', 'navy', 'ltGrey', 'sandyBrown', 'darkMagenta', 'mediumPurple', 'mediumSpringGreen', 'gainsboro', 'ghostWhite', 'salmon', 'medTurquoise', 'chocolate', 'ltPink', 'lightBlue', 'dkSalmon', 'white', 'indigo', 'ltYellow', 'sienna', 'mediumVioletRed', 'ltCyan', 'deepSkyBlue', 'springGreen', 'fuchsia', 'darkGray', 'darkSlateGrey', 'dkKhaki', 'burlyWood', 'ltSalmon', 'pink', 'darkSeaGreen', 'oldLace', 'red', 'darkGoldenrod', 'dkOliveGreen', 'lavenderBlush', 'greenYellow', 'lightSkyBlue', 'lightGoldenrodYellow', 'ltGreen', 'aquamarine', 'magenta', 'lightGreen', 'medOrchid', 'medPurple', 'medSeaGreen', 'medSpringGreen', 'blanchedAlmond', 'teal', 'powderBlue', 'orchid', 'turquoise', 'medVioletRed', 'lavender', 'cyan', 'oliveDrab', 'yellowGreen', 'cadetBlue', 'ivory', 'lightCyan', 'darkRed', 'saddleBrown', 'grey', 'lightSalmon', 'ltCoral', 'chartreuse', 'papayaWhip', 'skyBlue', 'dkCyan', 'forestGreen', 'brown', 'navajoWhite', 'blue', 'beige', 'dkGray', 'maroon', 'cornsilk', 'violet', 'mediumSeaGreen', 'deepPink', 'dkOrchid', 'orange', 'ltSeaGreen', 'medAquamarine', 'honeydew', 'mistyRose', 'mediumBlue', 'lightYellow', 'mediumAquamarine', 'slateGray', 'darkGreen', 'slateGrey', 'darkKhaki', 'gold', 'royalBlue', 'wheat', 'gray', 'aliceBlue', 'seaGreen', 'paleTurquoise', 'ltSlateGrey', 'dkOrange', 'ltSkyBlue', 'rosyBrown', 'silver', 'cornflowerBlue', 'azure', 'mediumSlateBlue', 'darkTurquoise', 'moccasin', 'lightSeaGreen', 'green', 'khaki', 'yellow', 'dimGray', 'lightPink', 'whiteSmoke', 'darkCyan', 'dkGrey', 'dkGreen', 'peru', 'darkOrchid', 'goldenrod', 'ltBlue', 'dkMagenta', 'dkSlateGray', 'snow', 'dkSlateBlue', 'darkSlateBlue', 'ltGray', 'lemonChiffon', 'dkRed', 'dimGrey', 'darkBlue', 'lightSteelBlue', 'ltGoldenrodYellow', 'thistle', 'darkSlateGray', 'plum', 'linen', 'seaShell', 'orangeRed', 'tan', 'dkTurquoise', 'indianRed', 'aqua', 'blueViolet', 'mediumOrchid','coral', 'midnightBlue', 'purple', 'paleVioletRed', 'lightSlateGrey','mintCream', 'medBlue', 'antiqueWhite' ]

class TableHandler():
    
    def __init__(self):
        print("This is constructor of ConvertTable")
        
    @staticmethod
    def getInitializedNumpyTable( numColumn, numRow, colNames=None ):
        if colNames==None:
            tab_np = np.zeros(numRow, dtype=[('V'+str(icol),'f8') for icol in range(numColumn)])
        else:
            tab_np = np.zeros(numRow, dtype=[(colNames[icol],'f8') for icol in range(numColumn)])
        return tab_np

    @staticmethod
    def getInitializedPandasTable( numColumn, numRow, colNames=None ):
        if colNames==None:
            tab_np = np.zeros(numRow, dtype=[('V'+str(icol),'f8') for icol in range(numColumn)])
        else:
            tab_np = np.zeros(numRow, dtype=[(colNames[icol],'f8') for icol in range(numColumn)])
        return pd.DataFrame(tab_np)
    
    @staticmethod
    def getNumpyTable(tab):
        if type(tab) is np.ndarray:
            tab_np = tab
        elif type(tab) is pd.core.frame.DataFrame:
            tab_np = np.zeros(tab.shape[0], dtype=[(tab.columns[i],'f8') for i in range(tab.shape[1])])
            for irow in range(tab.shape[0]):
                for icol in range(tab.shape[1]):
                    tab_np[irow][icol] = tab.values[irow][icol]
        else:
            tab_np = None
        return tab_np
    
    @staticmethod
    def getPandasTable(tab):
        if type(tab) is np.ndarray:
            tab_pd = pd.DataFrame(tab)
        elif type(tab) is pd.core.frame.DataFrame:
            tab_pd = tab
        else:
            tab_pd = None
        return tab_pd
    
    @staticmethod
    def getColumnDataByNames(tab,colNames):
        if type(tab) is np.ndarray:
            colData = tab[colNames]
        else:
            colData = None
        return colData
    
    @staticmethod
    def getColumnDataByIndex(tab,icolStart,icolEnd):
        if type(tab) is np.ndarray:
            colData = tab[list(tab.dtype.names[icolStart:(icolEnd+1)])]
        else:
            colData = None
        return colData
    
    @staticmethod
    def getSubarray(tab,icolStart,icolEnd,irowStart,irowEnd):
        if type(tab) is np.ndarray:
            subArray = tab[list(tab.dtype.names[icolStart:(icolEnd+1)])][irowStart:(irowEnd+1)]
        else:
            subArray = None
        return subArray

    @staticmethod
    def getColumnNumber(tab):
        if type(tab) is np.ndarray:
            return tab.shape[1]
        elif type(tab) is pd.core.frame.DataFrame:
            return len(tab.columns)

    @staticmethod
    def getRowNumber(tab):
        if type(tab) is np.ndarray:
            return tab.shape[0]
        elif type(tab) is pd.core.frame.DataFrame:
            return len(tab)

    @staticmethod
    def SaveAsExcel_TableSet(tableSet,filePath=None,sheetNameList=None):
        for i in range(len(tableSet)):
            tab    = tableSet[i]
            df     = TableHandler.getPandasTable(tab)
            if i==0:
                with pd.ExcelWriter(filePath, engine="openpyxl") as writer:
                    df.to_excel(writer, sheet_name=sheetNameList[i], index=False)
            else:
                with pd.ExcelWriter(filePath, engine="openpyxl", mode="a") as writer:
                    df.to_excel(writer, sheet_name=sheetNameList[i], index=False)

    @staticmethod 
    def addExcelPlotByAllSheets(fileNameExcel=None,indexY=None,indexX=None,yMin=None,yMax=None,indexSheetPlot=0,plotLocation="A1",plotHeight=None,plotWidth=None):
        varNameList    = TableHandler.getVariableNameList(fileNameExcel=fileNameExcel)
        sheetNameList  = TableHandler.getSheetNameList   (fileNameExcel=fileNameExcel)
        numVars        = len(sheetNameList)
        indexYList     = [indexY for ivr in range(numVars)]
        indexSheetList = [ivr for ivr in range(numVars)]
        xLabel         = varNameList[indexX-1]
        yLabel         = varNameList[indexY-1]

        for ivr in range(numVars):
            sheetNameList[ivr] = yLabel+" "+sheetNameList[ivr]

        TableHandler.OpenExcelAddPlot(filePath=fileNameExcel,indexSheetPlot=indexSheetPlot,indexSheetList=indexSheetList,indexX=indexX,indexYList=indexYList,
            colorList=[EXCEL_COLOR_LIST[ivr] for ivr in range(numVars)],legendList=sheetNameList,plotLocation=plotLocation,
            xLabel=xLabel,yLabel=yLabel,ymin=yMin,ymax=yMax,lineType=[""],plotHeight=plotHeight,plotWidth=plotWidth)

    @staticmethod
    def OpenExcelAddPlot(filePath=None,indexSheetPlot=None,indexX=1,indexSheetList=None,indexYList=None,colorList=None,legendList=None,lineType=None,plotLocation=None,xLabel=None,yLabel=None,ymin=None,ymax=None,plotHeight=None,plotWidth=None):
        wb       = openpyxl.load_workbook(filePath)
        ws       = wb.worksheets[indexSheetPlot]
        
        from openpyxl.chart import ScatterChart, Reference, Series
        from openpyxl.drawing.colors import ColorChoice
        from openpyxl.styles import colors
        
        chart      = ScatterChart(auto_axis=False)
        len_data   = ws.max_row-1
        num_series = len(indexYList)
        min_row    = 2
        max_row    = min_row + len_data - 1
        x_values   = Reference(ws, min_col=indexX, min_row = min_row, max_row = max_row)

        for i in range(0, num_series):
            min_col = indexYList[i]
            ws_dat  = wb.worksheets[indexSheetList[i]]
            values  = Reference(ws_dat, min_col = min_col, min_row = min_row, max_row = max_row)
            con     = Series(values, x_values, title=legendList[i])

            con.spPr.ln.solidFill = ColorChoice(prstClr=colorList[i])

            con.spPr.ln.prstDash  = None
            if lineType==None:
                con.spPr.ln.prstDash  = 'solid'

            con.spPr.ln.w = 1.5 * 12700

            con.marker.symbol = "circle"
            con.graphicalProperties.line.solidFill        = ColorChoice(prstClr=colorList[i])
            con.marker.graphicalProperties.solidFill      = ColorChoice(prstClr=colorList[i])
            con.marker.graphicalProperties.line.solidFill = ColorChoice(prstClr="black")
            con.marker.size = 5
            con.marker.spPr.noFill = False
            chart.series.append(con)

        chart.y_axis.scaling.min = ymin
        chart.y_axis.scaling.max = ymax
        chart.x_axis.title = xLabel
        chart.y_axis.title = yLabel

        if plotHeight!=None:
            chart.height = plotHeight

        if plotWidth!=None:
            chart.width  = plotWidth

        ws.add_chart(chart, plotLocation) 
        wb.save(filePath)

    @staticmethod
    def SaveAsExcel(tab,filePath=None,sheetName=None,withPlot=False,indexX=1,indexYList=None,colorList=None):
        if type(tab) is np.ndarray:
            tab_pd = pd.DataFrame(tab)
        elif type(tab) is pd.core.frame.DataFrame:
            tab_pd = tab

        wb       = openpyxl.Workbook()
        ws       = wb.active
        ws.title = sheetName

        stcol = 1
        strow = 1
        for c in range(0,len(tab_pd.columns)):
            ws[get_column_letter(c+stcol)+str(strow)].value = tab_pd.columns[c]
        
        numColumn = TableHandler.getColumnNumber(tab)
        numRow    = TableHandler.getRowNumber(tab)
        for r in range(0,numRow):
            for c in range(0,numColumn):
                ws[get_column_letter(c+stcol)+str(strow+r+1)].value = tab_pd.iloc[r][c]
        
        if withPlot==True:
            from openpyxl.chart import ScatterChart, Reference, Series
            from openpyxl.drawing.colors import ColorChoice
            from openpyxl.styles import colors
            
            chart      = ScatterChart()
            len_data   = len(tab_pd)
            num_series = len(indexYList)
            min_row    = 2
            max_row    = min_row + len_data - 1
            x_values   = Reference(ws, min_col=indexX, min_row = min_row, max_row = max_row)

            for i in range(0, num_series):
                min_col = indexYList[i]
                values  = Reference(ws, min_col = min_col, min_row = min_row, max_row = max_row)
                con     = Series(values, x_values, title=tab_pd.columns.values[min_col-1])
                
                if i==0:
                    con.spPr.ln.solidFill = ColorChoice(prstClr="red")
                elif i==1:
                    con.spPr.ln.solidFill = ColorChoice(prstClr="green")
                elif i==2:
                    con.spPr.ln.solidFill = ColorChoice(prstClr="blue")

                chart.series.append(con)
                
            ws.add_chart(chart, "N3") 

        wb.save(filePath)
    
    @staticmethod
    def addHeaderOfColumnNames(fileName=None,colNameList=None,colUnitList=None):

        if colNameList!=None and colUnitList==None:
            header = ""
            for i in range(len(colNameList)):
                if i!=len(colNameList)-1:
                    header += colNameList[i] + ","
                else:
                    header += colNameList[i]
            header += "\n"
        elif colNameList!=None and colUnitList!=None:
            header = ""
            for i in range(len(colNameList)):
                if i!=len(colNameList)-1:
                    header += colNameList[i] +"[" + colUnitList[i] + "],"
                else:
                    header += colNameList[i] +"[" + colUnitList[i] + "]"
            header += "\n"

        FileHandlerAscii.addLineOnTop(fileName=fileName,line=header)

    @staticmethod
    def getColumnTypeList(fileName=None,skiprows=None):
        tab = np.loadtxt(fname=fileName,skiprows=skiprows,delimiter=",")
        if type(tab) is np.ndarray:
            columnTypeList = []
            irow = 0
            for icol in range(tab.shape[1]):
                value = tab[irow][icol]
                columnTypeList.append(type(value))
            del tab
            return columnTypeList
        else:
            return None

    @staticmethod
    def openAsNumpyTable(fileName=None,skiprows=None,hasHeader=False):
        ext = FileHandlerAscii.getExtension(fileName=fileName)
        if ext=="csv":
            if hasHeader:
                colNameList = TableHandler.getHeaderAsColumnNameList(fileName=fileName)
                colTypeList = TableHandler.getColumnTypeList(fileName=fileName,skiprows=skiprows)
                if len(colNameList)!=len(colTypeList):
                    print('Error: length of colNameList and colTypeList should be the same', file=sys.stderr)
                    return None
                tab = np.loadtxt(fname=fileName,skiprows=skiprows,delimiter=",",dtype=[(colNameList[icol],colTypeList[icol]) for icol in range(len(colNameList))])
            return tab
        return None

    @staticmethod
    def readExcelDataAsPandasDfList(fileNameExcel=None):
        input_file       = pd.ExcelFile(fileNameExcel)
        input_sheet_name = input_file.sheet_names

        df_list=[]
        for sheet in input_sheet_name:
            df_list.append(input_file.parse(sheet))

        return df_list

    @staticmethod
    def readExcelDataAsPandasDfSingle(fileNameExcel=None):
        input_file       = pd.ExcelFile(fileNameExcel)
        input_sheet_name = input_file.sheet_names
        num_sheet        = len(input_sheet_name)

        dfList=[]
        for sheet in input_sheet_name:
            dfList.append(input_file.parse(sheet))

        df=pd.DataFrame()
        for i in range(num_sheet) :
            df=df.append(dfList[i])[dfList[0].columns.tolist()]
        df=df.dropna()

        return df

    @staticmethod
    def writePandasDataframeListAsExcel(dfList=None,fileNameExcel=None,sheetNameList=None):
        with pd.ExcelWriter(fileNameExcel,engine='openpyxl', mode='w') as writer:
            for i in range(len(dfList)):
                df = dfList[i]
                df = df.rename(columns={'Longigude[deg]':'Longitude[deg]'})
                print(df.columns)
                if sheetNameList!=None:
                    df.to_excel(writer, sheet_name=sheetNameList[i], index=False, encoding='utf-8')
                else:
                    df.to_excel(writer, sheet_name="Sheet"+str(i+1), index=False, encoding='utf-8')

    @staticmethod
    def getPandasDfListFromCsvs(fileNameListCsv=None):
        dfList        = []
        sheetNameList = []
        for fileNameCsv in fileNameListCsv:
            print(fileNameCsv)
            sheetNameList.append(FileHandlerAscii.getFileNameBase(fileName=fileNameCsv))
            tab = TableHandler.openAsNumpyTable(fileName=fileNameCsv,skiprows=1,hasHeader=True)
            df  = pd.DataFrame(tab)
            dfList.append(df)
        return dfList, sheetNameList

    @staticmethod
    def convertCsvFilesToExcel(fileNameListCsv=None,fileNameExcel=None,sheetNameList=None):
        dfList, sheetNameList_Loc = TableHandler.getPandasDfListFromCsvs(fileNameListCsv=fileNameListCsv)
        if sheetNameList==None:
            sheetNameList= sheetNameList_Loc
        TableHandler.writePandasDataframeListAsExcel(dfList,fileNameExcel=fileNameExcel,sheetNameList=sheetNameList)

    @staticmethod
    def getHeaderAsColumnNameList(fileName=None):
        ext = FileHandlerAscii.getExtension(fileName=fileName)
        if ext=="csv":
            f = open(fileName)
            header = f.readline()
            headerList = header.split(',')

            for i in range(len(headerList)):
                headerItem = headerList[i]
                headerList[i] = headerItem.replace("\n","")

            return headerList
        elif ext=="xlsx" or ext=="xls":
            tab = pd.read_excel(fileName,header=0)
            headerList = list(tab.columns)
            del tab
            return headerList

    @staticmethod
    def getUnitFromColumnName(columnName=None):
        columnName = re.split('\[',columnName)
        unitBuff   = re.split('\]',columnName[1])
        unit       = unitBuff[0]
        return unit

    @staticmethod
    def getNameFromColumnName(columnName=None):
        colName  = re.split('\[',columnName)
        name     = colName[0]
        return name

    @staticmethod
    def getUnitListFromColumnNameList(columnNameList):
        columnUnitList = []
        for i in range(len(columnNameList)):
            unit    = TableHandler.getUnitFromColumnName(columnName=columnNameList[i])
            columnUnitList.append(unit)
        return columnUnitList

    @staticmethod
    def getNameListFromColumnNameList(columnNameList):
        varNameList = []
        for i in range(len(columnNameList)):
            name = TableHandler.getNameFromColumnName(columnName=columnNameList[i])
            varNameList.append(name)
        return varNameList

    @staticmethod
    def setData(tab=None,icol=None,irow=None,value=None):
        if type(tab) is np.ndarray:
            tab[irow][icol] = value
        elif type(tab) is pd.core.frame.DataFrame:
            tab.iat[irow,icol] = value

    @staticmethod
    def setDataRow(tab=None,irow=None,valueList=None):
        for icol in range(len(valueList)):
            TableHandler.setData(tab,icol=icol,irow=irow,value=valueList[icol])
    
    @staticmethod
    def convertCsvToExcel(fileNameCsv=None,fileNameExcel=None,index=None):
        df = pd.read_csv(fileNameCsv)
        fileNameBase = FileHandlerAscii.getFileNameBase(fileName=fileNameCsv)
        if fileNameExcel==None:
            fileNameExcel = fileNameBase+".xlsx"
        df.to_excel(fileNameExcel,index=index,sheet_name=fileNameBase)

    @staticmethod
    def getSheetNameList(fileNameExcel=None):
        input_file = pd.ExcelFile(fileNameExcel)
        return input_file.sheet_names

    @staticmethod
    def getVariableNameList(fileNameExcel=None,isheet=0):
        df = pd.read_excel(fileNameExcel,header=0)
        varNameList = list(df.columns)
        del df
        return varNameList

if __name__ == '__main__':
    
    #_/_/_/_/ Initialize Numpy Array with Column Name
    # (1) Initialize Matrix
    tab = TableHandler.getInitializedNumpyTable( numColumn=3, numRow=10, colNames=['V1','V2','V3'] )
    
    # (2) Read from file
    #tab = np.genfromtxt( "test.csv", skip_header=1, delimiter=",", dtype=[('V1', 'f8'), ('V2', 'f8'), ('V3', 'f8')] )
    tab[0][0] = 1.0
    tab[1][0] = 0.5
    tab[2][0] = 0.25
    tab[1][1] = 2.0
    tab[2][2] = 3.0
    
    # Access column data by Column Name (NumpyTable)  [Done]
    colData = tab['V1']
    print(colData)
    
    # Access column data by Column Index (NumpyTable) [Done]
    icolSt  = 0
    icolEd  = 1
    colData = TableHandler.getColumnDataByIndex(tab,icolSt,icolEd)
    print("#### Access column data by column index")
    print(colData)
    
    # Access sub-array (NumpyTable) [Done]
    icolSt  = 0
    icolEd  = 1
    irowSt  = 0
    irowEd  = 3
    subArray = TableHandler.getSubarray(tab,icolSt,icolEd,irowSt,irowEd)
    print("#### Access sub-array")
    print(subArray)
    
    # Display Table Contents [----]
    # Display Table Column Names [----]
    print(tab)
    print(tab.dtype.names)
    
    #_/_/_/_/ Convert to Pandas DataFrame [Done]
    import pandas as pd
    # Convet to pandas table
    df = pd.DataFrame(tab)
    # Display Pandas Dataframe
    print(df)
    
    #_/_/_/_/ Convert Pandas DataFrame to Numpy Table [Done]
    tab_np = TableHandler.getNumpyTable(df)
    print("\n >>Converted to Numpy Table:\n")
    print(tab_np)
    print(tab_np.dtype.names)
    
    #_/_/_/_/ Convert Numpy Table to Pandas DataFrame [Done]
    tab_pd = TableHandler.getPandasTable(tab)
    print("\n >>Converted to Pandas Dataframe:\n")
    print(tab_pd)
